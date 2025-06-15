import asyncio
import os
import random
from telethon import TelegramClient, errors, functions, types
from telethon.errors import ChatSendMediaForbiddenError
from telethon.errors.rpcerrorlist import FloodWaitError
from openpyxl import load_workbook, Workbook
from datetime import datetime
from Text import FullText

# 🔐 API Telegram
api_id = 22325330
api_hash = '2cd1e042d95dfe4cf34b4d58309820cc'
session_name = 'main_accounti'

media_folder = 'media'
excel_file = 'result.xlsx'
pending_file = 'pending.xlsx'
log_file = 'log.txt'
monitor_file = 'sent_monitoring.xlsx'

# === Логирование ===
def log(message):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(f"[{ts}] {message}\n")

# === Чтение ссылок из Excel ===
def read_links_from_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active
    links = []
    for row in ws.iter_rows(min_row=3, min_col=7, max_col=7):
        cell = row[0]
        if cell.value and isinstance(cell.value, str) and 't.me' in cell.value:
            links.append((cell.value.strip(), cell))
    return wb, ws, links

# === Сохранение ссылки в pending.xlsx ===
def save_pending_link(link):
    try:
        wb_p = load_workbook(pending_file) if os.path.exists(pending_file) else Workbook()
        ws_p = wb_p.active
        if ws_p.max_row == 1 and ws_p.cell(1,1).value != 'Ссылка':
            ws_p.insert_rows(1)
            ws_p.cell(1,1,'Ссылка')
        ws_p.append([link])
        wb_p.save(pending_file)
        log(f"🕓 Ссылка добавлена в ожидание: {link}")
    except Exception as e:
        log(f"❌ Не удалось сохранить в pending.xlsx: {e}")

# === Сохранение отчёта мониторинга ===
def save_monitoring(records):
    try:
        wb_m = load_workbook(monitor_file) if os.path.exists(monitor_file) else Workbook()
        ws_m = wb_m.active
        # Устанавливаем заголовки
        if ws_m.max_row < 2:
            ws_m.delete_rows(1)
            ws_m.append(['Ссылка','Название','Дата','Статус'])
        for link, title, dt, status in records:
            ws_m.append([link, title, dt, status])
        wb_m.save(monitor_file)
        log(f"✅ Отчёт мониторинга сохранён: {len(records)} записей")
    except Exception as e:
        log(f"❌ Ошибка при сохранении отчёта мониторинга: {e}")

# === Отправка медиа и текста с пересылкой в 'Избранное' ===
async def send_media_message(client, entity, message, media_folder):
    title = getattr(entity, 'title', str(getattr(entity,'id','unknown')))
    images = [os.path.join(media_folder, f) for f in os.listdir(media_folder)
              if f.lower().endswith(('.jpg','.jpeg','.png'))]
    if not images:
        log(f"⚠️ В папке '{media_folder}' нет изображений")
        return False, title, None
    photo = random.choice(images)
    sent_media = None
    # Отправляем фото
    try:
        sent_media = await client.send_file(entity, photo)
        log(f"📸 Фото отправлено в: {title}")
    except ChatSendMediaForbiddenError:
        log(f"🚫 Отправка медиа запрещена для: {title}")
    except FloodWaitError as e:
        log(f"⏳ Задержка (FloodWait) при отправке медиа в {title}: ждем {e.seconds} сек")
        await asyncio.sleep(e.seconds+5)
        return False, title, None
    except Exception as e:
        log(f"❌ Ошибка при отправке фото в {title}: {e}")
    await asyncio.sleep(random.uniform(2,5))
    # Отправляем текст
    try:
        sent_text = await client.send_message(entity, message, parse_mode='HTML')
        log(f"✉️ Текст отправлен в: {title}")
        # Пересылаем в 'Избранное'
        try:
            to_me = 'me'
            if sent_media:
                await client.forward_messages(to_me, sent_media.id, entity)
            await client.forward_messages(to_me, sent_text.id, entity)
            log(f"📂 Переслано в 'Избранное' из: {title}")
        except Exception as e:
            log(f"⚠️ Не удалось переслать в 'Избранное' из {title}: {e}")
        return True, title, sent_text
    except errors.ChatWriteForbiddenError as e:
        log(f"🚫 Отправка текста запрещена для: {title}: {e}")
        return False, title, None
    except FloodWaitError as e:
        log(f"⏳ Задержка (FloodWait) при отправке текста в {title}: ждем {e.seconds} сек")
        await asyncio.sleep(e.seconds+5)
        return False, title, None
    except Exception as e:
        log(f"❌ Ошибка при отправке текста в {title}: {e}")
        return False, title, None

# === Основная логика ===
async def main():
    wb, ws, links_with_cells = read_links_from_excel(excel_file)
    monitor_records = []
    log("🚀 Скрипт запущен")
    client = TelegramClient(session_name, api_id, api_hash)
    await client.start()
    for link, cell in links_with_cells:
        try:
            log(f"➡️ Обработка ссылки: {link}")
            entity = await client.get_entity(link)
            # Вступление
            try:
                await client(functions.channels.JoinChannelRequest(channel=entity))
                log(f"✅ Вступили в: {link}")
            except FloodWaitError as e:
                log(f"⏳ Задержка (FloodWait) при вступлении в {link}: ждем {e.seconds} сек")
                await asyncio.sleep(e.seconds+5)
                try:
                    await client(functions.channels.JoinChannelRequest(channel=entity))
                    log(f"✅ После ожидания вступили в: {link}")
                except Exception as ee:
                    log(f"❌ Повторное вступление не удалось для {link}: {ee}")
                    save_pending_link(link)
                    cell.value = None
                    continue
            except errors.UserAlreadyParticipantError:
                log(f"🔹 Уже являетесь участником: {link}")
            except Exception as e:
                if 'requested to join' in str(e):
                    log(f"🕓 Запрос на вступление отправлен, ждем подтверждения: {link}")
                    save_pending_link(link)
                    cell.value = None
                    continue
                log(f"❌ Ошибка вступления в {link}: {e}")
                continue
            # Отправка
            success, title, sent_text = await send_media_message(client, entity, FullText, media_folder)
            status = 'Отправлено' if success else 'В ожидании'
            monitor_records.append((link, title, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), status))
            # Очистка и pending если неуспешно
            if not success:
                save_pending_link(link)
                try:
                    await client(functions.channels.LeaveChannelRequest(channel=entity))
                    log(f"🚪 Вышли из: {link}")
                except Exception as e:
                    log(f"⚠️ Ошибка при выходе из {link}: {e}")
            cell.value = None
            await asyncio.sleep(random.uniform(7,12))
        except Exception as e:
            log(f"❌ Общая ошибка при обработке {link}: {e}")
            continue
    save_monitoring(monitor_records)
    wb.save(excel_file)
    await client.disconnect()
    log("✅ Скрипт завершил работу")

if __name__ == '__main__':
    asyncio.run(main())
